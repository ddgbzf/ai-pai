# -*- coding: utf-8 -*-
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
import logging
from datetime import datetime
from openpyxl.utils import get_column_letter  # 用于列号转字母

class 排工配置:
    固定列颜色 = PatternFill(
        start_color="FFFFFF00",  # 预期开始色号（黄色，带透明度FF）
        end_color="FFFFFF00",    # 保留结束色号定义，但代码中仅检测开始色号
        fill_type="solid"
    )
    最大人员数 = 100        # 最大支持人员数（根据表格实际行数调整）

class 智能排工优化版:
    def __init__(self, 文件路径):
        self.文件路径 = 文件路径
        self.工序优先级 = {}  # 工序在评分表中的列索引（越小优先级越高）
        self._重置系统状态()
        self._配置日志输出()

    def _重置系统状态(self):
        """重置系统核心状态"""
        self.技能评分表 = None
        self.产品信息库 = {}
        self.固定岗位坐标 = {}  # 存储(人员列, 行号): 人员姓名
        self.请假人员列表 = []
        self.已分配人员集合 = set()  # 包含固定岗位和动态分配人员
        self.固定列标识 = set()
        self.当前批次编号 = datetime.now().strftime("%Y%m%d%H%M%S")
        self.全体人员 = []         # 全体有效人员列表
        self.总人数 = 0            # 总人数（工序评分表A列非空数-1）
        self.请假人数 = 0          # 请假人数（智能排工表A列非空数-1）
        self.总需求人数 = 0        # 总需求人数（主表第二行需求列之和）
        self.全局工序队列 = []  # 存储所有(优先级, 产品列, 行号, 工序, 产品名)
        self.产品列映射 = {}    # 存储产品名到列编号的映射
        self.动态状态列号 = None  # 动态计算的状态列位置

    def _配置日志输出(self):
        """初始化日志系统（终端输出）"""
        logging.basicConfig(
            level=logging.INFO,
            format='[%(asctime)s] %(message)s',
            datefmt='%Y-%m-%d %H:%M',        # 新增：日期格式精确到分钟
            handlers=[logging.StreamHandler()]
        )
        #logging.info(f"排产系统启动 - 批次ID：{self.当前批次编号}")

    def 执行排产流程(self):
        """主流程控制"""
        try:
            self._重置系统状态()
            工作簿, 主表 = self._加载基础数据()
            self._识别固定岗位(主表)         
            self._清除历史数据(主表)
            self._执行全局动态分配(主表)  
            self._计算总需求人数(主表)     
            self._生成状态报告(主表)
            
            工作簿.save(self.文件路径)
            logging.info("排产流程成功完成")
            return True
        except Exception as 错误信息:
            logging.error(f"排产失败：{str(错误信息)}", exc_info=True)
            return False

    def _加载基础数据(self):
        """加载Excel中的核心数据（恢复工序流程表需求加载）"""
        # 加载技能评分表并建立工序优先级（列索引越小优先级越高）
        try:
            self.技能评分表 = pd.read_excel(
                self.文件路径,
                sheet_name='工序评分',
                index_col=0,
                engine='openpyxl'
            ).fillna(0)
            self.工序优先级 = {col: idx for idx, col in enumerate(self.技能评分表.columns)}
        except KeyError:
            raise Exception("缺失'工序评分'工作表，请检查文件结构")

        # 加载工序流程表（包含需求人数，作为源头数据）
        try:
            产品数据表 = pd.read_excel(
                self.文件路径,
                sheet_name='工序流程',
                engine='openpyxl'
            )
            工序列筛选 = [col for col in 产品数据表.columns if col.startswith('工序')]
            
            for _, 行数据 in 产品数据表.iterrows():
                产品名称 = 行数据.iloc[0]          # 按位置取第一列（产品名称）
                产品产能 = 行数据.get('产能', 0)
                需求人数 = 行数据.get('人数', 0)     # 从工序流程表获取需求人数
                工序步骤 = [
                    str(行数据[col]).strip() 
                    for col in 工序列筛选 
                    if not pd.isna(行数据[col])
                ]
                self.产品信息库[产品名称] = {
                    '产能': 产品产能,
                    '需求人数': 需求人数,  # 保留工序流程表的需求数据
                    '工序': 工序步骤 
                }
        except Exception as 错误:
            raise Exception(f"工序流程表加载失败：{str(错误)}")

        # 加载主表并建立产品列映射（产品名到列编号）
        try:
            工作簿 = openpyxl.load_workbook(self.文件路径)
            主表 = 工作簿['智能排工']
            
            # 计算总人数（工序评分表A列非空单元格数，从第2行开始）
            评分表 = 工作簿['工序评分']
            self.全体人员 = [
                self._清洗姓名(row[0].value) 
                for row in 评分表.iter_rows(min_row=2, min_col=1) 
                if self._清洗姓名(row[0].value) is not None
            ]
            self.总人数 = len(self.全体人员)

            # 计算请假人数（智能排工表A列非空单元格数，从第2行开始）
            self.请假人员列表 = [
                self._清洗姓名(cell.value) 
                for cell in 主表['A'][1:排工配置.最大人员数 + 1] 
                if cell.value is not None
            ]
            self.请假人数 = len(self.请假人员列表)

            # 构建产品列映射
            self.产品列映射 = {}
            最大人员列 = 0  # 用于找到最后一个人员列
            
            for col in range(2, 主表.max_column + 1, 2): 
                产品名 = 主表.cell(1, col).value
                if 产品名:
                    self.产品列映射[产品名] = col
                    最大人员列 = max(最大人员列, col + 1)  # col+1是人员列
            
            # 计算动态状态列号（最后一个人员列后面第1列）
            if 最大人员列 > 0:
                self.动态状态列号 = 最大人员列 + 1
                #logging.info(f"检测到{len(self.产品列映射)}个产品，状态列自动设置在{get_column_letter(self.动态状态列号)}列")
            else:
                self.动态状态列号 = 12  # 默认值，如果没有产品列
                logging.warning("未检测到产品列，使用默认状态列位置")

            # 构建全局工序队列（包含所有产品的所有工序及其坐标）
            self.全局工序队列 = []
            for 产品名, 详情 in self.产品信息库.items():
                产品列 = self.产品列映射.get(产品名)
                if not 产品列:
                    continue
                for 行偏移, 工序 in enumerate(详情['工序'], start=3):
                    优先级 = self.工序优先级.get(工序, float('inf'))  # 不存在的工序优先级最低
                    self.全局工序队列.append((优先级, 产品列, 行偏移, 工序, 产品名))

            # 修改排序逻辑：先按优先级，再按产品列，最后按行号
            self.全局工序队列.sort(key=lambda x: (x[0], x[1], x[2]))

            return 工作簿, 主表
        except KeyError:
            raise Exception("缺失'智能排工'主表，请检查文件结构")

    def _清除历史数据(self, 主表):
        """清空主表中除标题行外的历史数据"""
        if self.动态状态列号 is None:
            raise Exception("状态列未初始化")
            
        for 行对象 in 主表.iter_rows(
            min_row=2, 
            max_row=主表.max_row, 
            min_col=2, 
            max_col=主表.max_column
        ):
            for 单元格 in 行对象:
                # 不清除动态状态列的内容
                if 单元格.column == self.动态状态列号:
                    continue
                单元格.value = None  # 清空所有数据单元格

    def _识别固定岗位(self, 主表):
        """提取标记为固定列的岗位及预分配人员（新增请假冲突检测）"""
        self.固定岗位坐标 = {}  # 重置固定岗位信息
        固定列信息 = []  # 记录固定列信息用于日志输出
        
        for 列对象 in 主表.iter_cols(min_row=1, max_row=1):
            标题单元格 = 列对象[0]
            if self._is_fixed_column(标题单元格):
                列号 = 标题单元格.column
                列字母 = get_column_letter(列号)
                人员列 = 列号 + 1  # 固定列右侧列是人员分配列
                固定人数 = 0
                跳过人数 = 0
                
                for 行号 in range(3, 排工配置.最大人员数 + 2):  # 人员从第3行开始
                    单元格 = 主表.cell(row=行号, column=人员列)
                    人员姓名 = self._清洗姓名(单元格.value)
                    if 人员姓名 and 人员姓名 in self.技能评分表.index:
                        if 人员姓名 in self.请假人员列表:
                            跳过人数 += 1  # 跳过请假的固定人员
                            continue
                        self.固定岗位坐标[(人员列, 行号)] = 人员姓名
                        self.已分配人员集合.add(人员姓名)
                        固定人数 += 1
                
                # 记录固定列信息
                if 固定人数 > 0 or 跳过人数 > 0:
                    产品名 = 主表.cell(1, 列号).value or f"第{列字母}列"
                    固定列信息.append(f"{产品名}({列字母}列)")
        
        # 输出固定列检测结果
        if 固定列信息:
            logging.info(f"检测到固定列：{', '.join(固定列信息)}")
        else:
            logging.info("未检测到固定列")
        
        logging.info(f"总人数：{self.总人数} 人")
        logging.info(f"请假人数：{self.请假人数} 人")

    def _is_fixed_column(self, 单元格):
        """判断是否为固定列（仅检测开始色号，移除调试日志）"""
        return (单元格.fill.start_color.index if 单元格.fill else "") == 排工配置.固定列颜色.start_color.index

    def _执行全局动态分配(self, 主表):
        """按全局工序优先级顺序分配人员（同步工序流程表需求到主表）"""
        可用人员 = list(set(self.全体人员) - set(self.请假人员列表) - self.已分配人员集合)

        # 写入产品基础数据（产能和需求人数，需求来自工序流程表）
        for 产品名, 详情 in self.产品信息库.items():
            产品列 = self._find_product_column(主表, 产品名)
            if not 产品列:
                continue
            主表.cell(2, 产品列, value=详情['产能'])
            主表.cell(2, 产品列 + 1, value=详情['需求人数'])  # 将工序流程需求写入主表

        # 按全局优先级处理每个工序
        for 优先级, 产品列, 行号, 工序, 产品名 in self.全局工序队列:
            工序单元格 = 主表.cell(row=行号, column=产品列)
            工序单元格.value = 工序
            self._分配具体岗位(主表, 产品列, 行号, 工序, 可用人员)

            # 动态更新可用人员
            可用人员 = list(set(self.全体人员) - set(self.请假人员列表) - self.已分配人员集合)

    def _find_product_column(self, 主表, 产品名):
        """根据产品名查找主表中的列编号"""
        return self.产品列映射.get(产品名)

    def _分配具体岗位(self, 主表, 产品列, 行号, 工序, 候选人池):
        """为工序分配人员（固定列优先 + 动态分配）"""
        人员列 = 产品列 + 1
        目标单元格 = 主表.cell(row=行号, column=人员列)
        坐标 = (人员列, 行号)

        # 处理固定岗位（优先分配，已排除请假人员）
        if 坐标 in self.固定岗位坐标:
            预分配人员 = self.固定岗位坐标[坐标]
            目标单元格.value = 预分配人员
            self.已分配人员集合.add(预分配人员)
            return  # 固定岗位分配后直接返回，无需动态分配

        # 非固定岗位动态分配
        合格候选人 = []
        for 人员姓名 in 候选人池:
            if 人员姓名 in self.已分配人员集合:
                continue
            try:
                技能评分 = self.技能评分表.loc[人员姓名, 工序]
                if 技能评分 > 0:
                    合格候选人.append((人员姓名, 技能评分))
            except KeyError:
                pass

        if 合格候选人:
            # 按评分从高到低排序
            合格候选人.sort(key=lambda x: x[1], reverse=True)
            最佳人员 = 合格候选人[0][0]
            目标单元格.value = 最佳人员
            self.已分配人员集合.add(最佳人员)

    def _计算总需求人数(self, 主表):
        """统计主表中已填充的需求列（确保工序流程需求已同步到主表）"""
        self.总需求人数 = 0
        for 列编号 in range(2, 主表.max_column + 1, 2):  # 产品列从第2列开始，间隔2列
            需求列 = 列编号 + 1
            需求值 = 主表.cell(2, 需求列).value or 0
            self.总需求人数 += 需求值
        logging.info(f"总需求人数：{int(self.总需求人数)} 人")

    def _写入状态列避开合并单元格(self, 主表, 行号, 列号, 值):
        """安全地写入状态列，避开合并单元格"""
        单元格 = 主表.cell(row=行号, column=列号)
        
        # 检查单元格是否是合并单元格的一部分
        for 合并范围 in 主表.merged_cells.ranges:
            if 单元格.coordinate in 合并范围:
                # 如果是合并单元格的一部分，找到合并范围的左上角单元格
                左上角 = 合并范围.min_row, 合并范围.min_col
                if (行号, 列号) != 左上角:
                    # 如果不是左上角单元格，尝试写入左上角单元格
                    左上角单元格 = 主表.cell(row=左上角[0], column=左上角[1])
                    if 左上角单元格.value is None or 左上角单元格.value == 值:
                        # 如果左上角单元格为空或与要写入的值相同，可以写入
                        左上角单元格.value = 值
                        return True
                    else:
                        # 如果左上角单元格已有不同值，可能需要处理冲突
                        logging.warning(f"合并单元格{单元格.coordinate}的左上角已有内容：{左上角单元格.value}")
                        return False
        
        # 如果不是合并单元格的一部分，正常写入
        单元格.value = 值
        return True

    def _生成状态报告(self, 主表):
        """生成状态报告"""
        if self.动态状态列号 is None:
            raise Exception("状态列未初始化")
            
        # 计算未分配人员：全体人员 - 请假人员 - 已分配人员
        未分配人员 = [
            name for name in self.全体人员
            if name not in self.请假人员列表 and name not in self.已分配人员集合
        ]
        
        # 计算状态描述（总可用人数 = 总人数 - 请假人数）
        可用人数 = self.总人数 - self.请假人数
        差值 = 可用人数 - self.总需求人数
        
        if 差值 > 0:
            logging.info(f"剩余人数：{int(差值)} 人")
        else:
            logging.info(f"欠缺人数：{abs(int(差值))} 人")
        
        # 写入状态描述（从第2行开始）
        状态列字母 = get_column_letter(self.动态状态列号)
        
        # 第2行：显示状态描述
        self._写入状态列避开合并单元格(主表, 2, self.动态状态列号, f"{'剩余' if 差值 >0 else '欠缺'}{int(abs(差值))}人")
        
        # 清空状态列第3行及以下的单元格（避免历史数据残留）
        for 行号 in range(3, 主表.max_row + 1):
            self._写入状态列避开合并单元格(主表, 行号, self.动态状态列号, None)
        
        # 填充未分配人员名单（从第3行开始）
        起始行 = 3
        for 索引, 人员姓名 in enumerate(未分配人员):
            行号 = 起始行 + 索引
            if 行号 <= 主表.max_row:
                self._写入状态列避开合并单元格(主表, 行号, self.动态状态列号, 人员姓名)
            
        logging.info(f"状态报告已生成在{状态列字母}列")

    @staticmethod
    def _清洗姓名(原始姓名):
        """清洗人员姓名（去除无效字符）"""
        if not 原始姓名:
            return None
        try:
            清洗后 = str(原始姓名).strip()
            return 清洗后 if 2 <= len(清洗后) <= 20 else None
        except:
            return None

if __name__ == "__main__":
    # 请根据实际路径修改文件路径（必须指向 .xlsx 文件，而非 .py 脚本）
    文件路径 = "/storage/emulated/0/Documents/智能排工系统.xlsx"
    排工系统 = 智能排工优化版(文件路径)
    if 排工系统.执行排产流程():
        print("运行结束")
    else:
        print("排产失败，请查看上方日志信息")